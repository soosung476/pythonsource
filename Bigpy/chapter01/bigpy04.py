import openpyxl
from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR.parent / "Py_Scrap" / "data"


# 워크북 인스턴스 객체 => excel 파일 생성
wb = openpyxl.Workbook()

# 활성화된 워크북에 워크시트 객체 => 시트 만들기
sheet = wb.active

# sheet 이름
sheet.title = '회원정보'

# header 컬럼만들기
header_titles = ['아이디', '전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1,column = idx+1, value = title)

# 내용 저장
members = [('happy','010-1234-5678'), ('smile','010-9876-5432')]

row_num = 2

for r,member in enumerate(members):
    for c, v in enumerate(member):
        sheet.cell(row = row_num, column = c+1, value = v)
    row_num += 1

wb.save(DATA_DIR / "members.xlsx")    
wb.close()