import openpyxl
from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR.parent / "Py_Scrap" / "data"

'''
wb = openpyxl.load_workbook(DATA_DIR / "sample.xlsx")
sheet = wb['Sheet1']
print(sheet.max_column, sheet.max_row)
print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=2, column=1).value)
wb.close()


# row 단위로 로딩

wb = openpyxl.load_workbook(DATA_DIR / "sample.xlsx")
sheet = wb['Sheet1']
cells = sheet['A2:C4']
for row in cells:
    for cell in row:
        print(cell.value)
    print('-'*10)
wb.close()
'''

# 전체 로딩
wb = openpyxl.load_workbook(DATA_DIR / "sample.xlsx")
sheet = wb['Sheet1']

# iter_rows() 함수를 이용해서 워크시트 내의 모든 row를 탐색
for row in sheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value)
    print('-'*10)
wb.close()